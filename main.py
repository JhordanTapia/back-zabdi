from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Form
from sqlalchemy.orm import Session
from pydantic import BaseModel
import bcrypt
import jwt
from datetime import datetime, timedelta, timezone
from database import get_db, engine, Base
from models import Usuario, Perfil, Presupuesto
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import os
import shutil
import requests
from extractor_ia import obtener_datos_json
from guardar_presupuesto import guardar_datos_revisados_en_bd
from fastapi.responses import StreamingResponse
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as ExcelImage
from pydantic import BaseModel
from dotenv import load_dotenv  # <-- 1. Importar la librería
from typing import Optional


load_dotenv()

Base.metadata.create_all(bind=engine)
app = FastAPI(title="API Presupuestos Navales")

# --- CONFIGURACIÓN CORS ---
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- CONFIGURACIÓN JWT ---
SECRET_KEY = os.getenv("SECRET_KEY")  # <-- 3. Leer desde el .env
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 1440  # 24 horas (1 día)


def crear_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


# --- ESQUEMAS ---
class LoginRequest(BaseModel):
    email: str
    password: str


class RegistroRequest(BaseModel):
    email: str
    password: str


class ItemEditado(BaseModel):
    detalle_actividad: str
    cantidad: float
    precio_unitario: float
    notas: str | None = None # <--- AÑADIDO


class PresupuestoGuardarRequest(BaseModel):
    cliente: str | None = None
    embarcacion: str
    numero_cotizacion: str | None = None
    fecha: str | None = None
    moneda: str = "PEN"  # <--- MAGIA MULTIMONEDA
    notas: str | None = None # <--- AÑADIDO
    items: list[ItemEditado]
    password_confirmacion: str

class ItemActualizar(BaseModel):
    id: int | None = None
    descripcion_original: str
    cantidad: float
    precio_unitario: float
    notas: str | None = None # <--- AÑADIDO

class PresupuestoActualizar(BaseModel):
    nro_cotizacion: str | None = None
    notas: str | None = None # <--- AÑADIDO
    items: list[ItemActualizar]


class EstadoActualizar(BaseModel):
    id_estado: int

class RegularizarProveedorRequest(BaseModel):
    ruc: str
    razon_social: str
    direccion: str = None
    distrito: str = None
    provincia: str = None
    departamento: str = None


# --- EL PORTERO DE SEGURIDAD ---
security = HTTPBearer()


def obtener_usuario_actual(credentials: HTTPAuthorizationCredentials = Depends(security),
                           db: Session = Depends(get_db)):
    token = credentials.credentials
    try:
        # Abrimos el token para ver si es válido
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("email")
        if email is None:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token inválido")

    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Tu token ya expiró")
    except jwt.PyJWTError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token falso o corrupto")

    # Verificamos que el usuario exista
    usuario = db.query(Usuario).filter(Usuario.email == email).first()
    if usuario is None:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="El usuario ya no existe")

    return usuario


# ==========================================================
#                    RUTAS DE USUARIOS
# ==========================================================

@app.post("/api/login")
def login(request: LoginRequest, db: Session = Depends(get_db)):
    usuario = db.query(Usuario).filter(Usuario.email == request.email).first()

    # Validar usuario y contraseña
    if not usuario:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Credenciales incorrectas")

    if not bcrypt.checkpw(request.password.encode('utf-8'), usuario.password_hash.encode('utf-8')):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Credenciales incorrectas")

    # Generar el Token
    token_jwt = crear_token({"sub": str(usuario.id), "email": usuario.email})

    return {
        "mensaje": "Login exitoso",
        "access_token": token_jwt,
        "token_type": "bearer",
        "usuario_id": str(usuario.id),
        "email": usuario.email
    }


@app.post("/api/registro")
def registrar_usuario(request: RegistroRequest, db: Session = Depends(get_db)):
    usuario_existente = db.query(Usuario).filter(Usuario.email == request.email).first()
    if usuario_existente:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ese correo ya está registrado")

    password_bytes = request.password.encode('utf-8')
    salt = bcrypt.gensalt()
    hashed_password = bcrypt.hashpw(password_bytes, salt).decode('utf-8')

    nuevo_usuario = Usuario(email=request.email, password_hash=hashed_password)
    db.add(nuevo_usuario)
    db.flush()  # flush() guarda temporalmente y nos da el ID para amarrarlo al perfil

    # Creamos su perfil asociado automáticamente
    nuevo_perfil = Perfil(
        id=nuevo_usuario.id,
        nombre_completo="Nuevo Operador",
        rol="Operador"
    )
    db.add(nuevo_perfil)

    db.commit()  # Ahora sí, guardamos el usuario y el perfil de golpe
    db.refresh(nuevo_usuario)

    return {"mensaje": "Usuario y Perfil creados exitosamente", "usuario_id": str(nuevo_usuario.id)}


@app.get("/api/protegido")
def ruta_privada(usuario_actual: Usuario = Depends(obtener_usuario_actual)):
    return {
        "mensaje": "¡Pasaste al sistema naval!",
        "usuario": usuario_actual.email
    }


# ==========================================================
#             NUEVA RUTA: PROCESADOR DE EXCELS
# ==========================================================

# Creamos una carpeta temporal si no existe para guardar los Excels
os.makedirs("temp", exist_ok=True)


@app.post("/api/presupuestos/analizar-excel")
async def analizar_excel_ia(
        archivo: UploadFile = File(...),
        moneda: str = Form(...), # <--- Atrapamos la moneda de Angular
        usuario_actual: Usuario = Depends(obtener_usuario_actual)
):
    if not archivo.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Mándame un archivo Excel válido (.xlsx o .xls)")

    ruta_temp = f"temp/{archivo.filename}"

    try:
        with open(ruta_temp, "wb") as buffer:
            shutil.copyfileobj(archivo.file, buffer)

        # Le pasamos la moneda humana a la función de la IA
        datos_ia = obtener_datos_json(ruta_temp, moneda)

        os.remove(ruta_temp)

        # 1. Capturamos si la IA devolvió la bandera de alto tráfico
        if isinstance(datos_ia, dict) and datos_ia.get("error_ia_trafico"):
            raise HTTPException(
                status_code=429,
                detail="Los servidores de IA están experimentando alto tráfico. Por favor, espera unos segundos y vuelve a intentarlo."
            )

        # 2. Capturamos errores generales si viene vacío
        if not datos_ia:
            raise HTTPException(status_code=500, detail="La IA falló al extraer los datos.")

        return {
            "status": "success",
            "mensaje": "Excel analizado. Esperando confirmación del usuario.",
            "data": datos_ia  # Mandamos el JSON al frontend para que lo vea en el Modal
        }



    except HTTPException as http_exc:

        # Si el error es nuestro querido 429 de tráfico (u otro error HTTP controlado),

        # lo dejamos pasar limpio hacia Angular sin convertirlo en 500.

        raise http_exc


    except Exception as e:

        # Para cualquier otra falla, intentamos borrar la basura y lanzamos el 500

        try:

            if os.path.exists(ruta_temp):
                os.remove(ruta_temp)

        except Exception:

            pass

        raise HTTPException(status_code=500, detail=f"Error analizando el documento: {str(e)}")


# ==========================================================
#         FASE 2: GUARDAR EN BD TRAS CONFIRMAR CON CLAVE
# ==========================================================
@app.post("/api/presupuestos/guardar-confirmado")
def guardar_presupuesto_confirmado(
        request: PresupuestoGuardarRequest,
        db: Session = Depends(get_db),
        usuario_actual: Usuario = Depends(obtener_usuario_actual)
):
    # 1. Verificar la contraseña del usuario logueado
    if not bcrypt.checkpw(request.password_confirmacion.encode('utf-8'), usuario_actual.password_hash.encode('utf-8')):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Contraseña incorrecta. Operación cancelada por seguridad."
        )

    # 2. Transformar el modelo de Pydantic a un diccionario de Python como lo espera nuestra función
    data_ia = request.model_dump()

    # 3. Guardar en Base de Datos usando la nueva función puente
    exito, mensaje = guardar_datos_revisados_en_bd(data_ia, usuario_actual.id, db)

    if not exito:
        raise HTTPException(status_code=500, detail=mensaje)

    return {"status": "success", "mensaje": mensaje}


# ==========================================================
#             NUEVA RUTA: LISTAR PRESUPUESTOS
# ==========================================================
@app.get("/api/presupuestos/lista")
def listar_presupuestos(
        papelera: bool = False,
        db: Session = Depends(get_db),
        usuario_actual: Usuario = Depends(obtener_usuario_actual)
):
    try:
        # Filtramos por la columna eliminado
        presupuestos = db.query(Presupuesto).filter(Presupuesto.eliminado == papelera).order_by(
            Presupuesto.fecha_emision.desc()).all()

        resultado = []
        for p in presupuestos:
            nombre_estado = p.estado_actual.nombre if p.estado_actual else "Borrador"
            nombre_astillero = p.proveedor.razon_social if p.proveedor else "Particular / S.E."

            resultado.append({
                "id": p.id,
                "numero_cotizacion": p.nro_presupuesto,
                "astillero": nombre_astillero,
                "embarcacion": p.proyecto_embarcacion,
                "fecha": p.fecha_emision.strftime("%Y-%m-%d") if p.fecha_emision else "Sin fecha",
                "estado": nombre_estado,
                "moneda": p.moneda if p.moneda else "PEN",
                "notas": p.notas
            })

        return {"status": "success", "data": resultado}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al extraer presupuestos de la BD: {str(e)}")


# --- NUEVAS RUTAS PARA SOFT DELETE Y RESTAURAR ---
@app.put("/api/presupuestos/{id_presupuesto}/soft-delete")
def enviar_a_papelera(
        id_presupuesto: int,
        db: Session = Depends(get_db),
        usuario_actual: Usuario = Depends(obtener_usuario_actual)
):
    presupuesto = db.query(Presupuesto).filter(Presupuesto.id == id_presupuesto).first()
    if not presupuesto:
        raise HTTPException(status_code=404, detail="No encontrado")

    presupuesto.eliminado = True
    db.commit()
    return {"status": "success"}


@app.put("/api/presupuestos/{id_presupuesto}/restaurar")
def restaurar_presupuesto(
        id_presupuesto: int,
        db: Session = Depends(get_db),
        usuario_actual: Usuario = Depends(obtener_usuario_actual)
):
    presupuesto = db.query(Presupuesto).filter(Presupuesto.id == id_presupuesto).first()
    if not presupuesto:
        raise HTTPException(status_code=404, detail="No encontrado")

    presupuesto.eliminado = False
    db.commit()
    return {"status": "success"}
@app.get("/api/presupuestos/{id_presupuesto}/detalle")
def obtener_detalle_presupuesto(
        id_presupuesto: int,
        db: Session = Depends(get_db),
        usuario_actual: Usuario = Depends(obtener_usuario_actual)
):
    try:
        from models import ItemPresupuesto

        # Obtenemos los ítems de ese presupuesto
        items = db.query(ItemPresupuesto).filter(ItemPresupuesto.id_presupuesto == id_presupuesto).all()

        resultado_items = []
        total_presupuesto = 0.0

        for item in items:
            subtotal = float(item.cantidad) * float(item.precio_unitario)
            total_presupuesto += subtotal

            resultado_items.append({
                "id": item.id,
                "descripcion": item.descripcion_original,
                "cantidad": float(item.cantidad),
                "precio_unitario": float(item.precio_unitario),
                "subtotal": subtotal,
                "notas": item.notas  # <--- ENVIAMOS LAS NOTAS DEL ÍTEM A ANGULAR
            })

        return {
            "status": "success",
            "data": {
                "items": resultado_items,
                "total": total_presupuesto
            }
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al extraer detalle: {str(e)}")


# ==========================================================
#             NUEVA RUTA: ACTUALIZAR PRESUPUESTO (CON DELETE)
# ==========================================================
@app.put("/api/presupuestos/{id_presupuesto}/actualizar")
def actualizar_presupuesto(
        id_presupuesto: int,
        request: PresupuestoActualizar,
        db: Session = Depends(get_db),
        usuario_actual: Usuario = Depends(obtener_usuario_actual)
):
    try:
        from models import ItemPresupuesto, Presupuesto

        # 1. Actualizamos la cabecera
        presupuesto_db = db.query(Presupuesto).filter(Presupuesto.id == id_presupuesto).first()
        if not presupuesto_db:
            raise HTTPException(status_code=404, detail="Presupuesto no encontrado")

        presupuesto_db.nro_presupuesto = request.nro_cotizacion
        presupuesto_db.notas = request.notas

        # 2. Capturamos los IDs de los ítems que sobrevivieron en el frontend (que SÍ tienen ID)
        ids_recibidos = [item.id for item in request.items if item.id is not None]

        # 3. ELIMINAR los que están en la BD pero ya no en la lista
        if ids_recibidos:
            db.query(ItemPresupuesto).filter(
                ItemPresupuesto.id_presupuesto == id_presupuesto,
                ~ItemPresupuesto.id.in_(ids_recibidos)
            ).delete(synchronize_session=False)
        else:
            # Si borraste ABSOLUTAMENTE TODOS los ítems
            db.query(ItemPresupuesto).filter(ItemPresupuesto.id_presupuesto == id_presupuesto).delete()

        # 4. ACTUALIZAR los que se quedaron o CREAR los nuevos añadidos
        for item_req in request.items:
            if item_req.id:
                # Si tiene ID, lo actualizamos
                item_db = db.query(ItemPresupuesto).filter(ItemPresupuesto.id == item_req.id).first()
                if item_db:
                    item_db.descripcion_original = item_req.descripcion_original
                    item_db.cantidad = item_req.cantidad
                    item_db.precio_unitario = item_req.precio_unitario
                    item_db.notas = item_req.notas  # <--- ACTUALIZA LA NOTA DEL ÍTEM
            else:
                # Si NO tiene ID, es una fila nueva que añadieron en el modal
                nuevo_item = ItemPresupuesto(
                    id_presupuesto=id_presupuesto,
                    descripcion_original=item_req.descripcion_original,
                    cantidad=item_req.cantidad,
                    precio_unitario=item_req.precio_unitario,
                    notas=item_req.notas  # <--- GUARDA LA NOTA DEL ÍTEM NUEVO
                )
                db.add(nuevo_item)

        db.commit()
        return {"status": "success", "mensaje": "Presupuesto actualizado correctamente"}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al actualizar: {str(e)}")


# ==========================================================
#             NUEVA RUTA: CAMBIAR ESTADO
# ==========================================================
@app.put("/api/presupuestos/{id_presupuesto}/estado")
def actualizar_estado_presupuesto(
        id_presupuesto: int,
        request: EstadoActualizar,
        db: Session = Depends(get_db),
        usuario_actual: Usuario = Depends(obtener_usuario_actual)
):
    try:
        presupuesto_db = db.query(Presupuesto).filter(Presupuesto.id == id_presupuesto).first()
        if not presupuesto_db:
            raise HTTPException(status_code=404, detail="Presupuesto no encontrado")

        # Actualizamos solo la columna del estado
        presupuesto_db.id_estado = request.id_estado
        db.commit()

        return {"status": "success", "mensaje": "Estado actualizado correctamente"}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al cambiar estado: {str(e)}")


# ==========================================================
#             NUEVA RUTA: CREACIÓN MANUAL DE PRESUPUESTO
# ==========================================================
class PresupuestoCrearManual(BaseModel):
    astillero: str | None = None
    embarcacion: str
    numero_cotizacion: str | None = None
    moneda: str = "PEN"
    notas: str | None = None # <--- AÑADIDO
    items: list[ItemEditado]


@app.post("/api/presupuestos/crear-manual")
def crear_presupuesto_manual(
        request: PresupuestoCrearManual,
        db: Session = Depends(get_db),
        usuario_actual: Usuario = Depends(obtener_usuario_actual)
):
    try:
        # Asegúrate de importar Proveedor o como se llame tu modelo de la tabla proveedores
        from models import Presupuesto, ItemPresupuesto, Proveedor
        from datetime import datetime

        proveedor_id = None

        # 1. LÓGICA DE PROVEEDOR (ASTILLERO) - NORMALIZADO A MAYÚSCULAS
        if request.astillero and request.astillero.strip():
            # El .upper() convierte mágicamente "sima" o "Sima" a "SIMA"
            texto_astillero = request.astillero.strip().upper()

            # Buscamos si ya existe el proveedor (comparación exacta porque ya todo está en mayúsculas)
            proveedor_db = db.query(Proveedor).filter(Proveedor.razon_social == texto_astillero).first()

            if proveedor_db:
                # Si ya existe, usamos su ID
                proveedor_id = proveedor_db.id
            else:
                # Si es nuevo, lo creamos como FANTASMA (sin RUC)
                nuevo_proveedor = Proveedor(
                    razon_social=texto_astillero,
                    nombre_comercial=texto_astillero
                )
                db.add(nuevo_proveedor)
                db.flush()  # Guardamos temporalmente para generar el ID
                proveedor_id = nuevo_proveedor.id

        # 2. CREAR CABECERA (Estado 1 = Borrador)
        nuevo_presupuesto = Presupuesto(
            creado_por=usuario_actual.id,
            id_proveedor=proveedor_id,
            proyecto_embarcacion=request.embarcacion,
            nro_presupuesto=request.numero_cotizacion,
            fecha_emision=datetime.now().date(),
            moneda=request.moneda,
            notas=request.notas,  # <--- GUARDAMOS LA NOTA NUEVA EN LA BD
            id_estado=1
        )
        db.add(nuevo_presupuesto)
        db.flush()  # Guardamos temporalmente para obtener el ID del presupuesto

        # 3. AGREGAR LOS ÍTEMS
        for req_item in request.items:
            nuevo_item = ItemPresupuesto(
                id_presupuesto=nuevo_presupuesto.id,
                descripcion_original=req_item.detalle_actividad,
                cantidad=req_item.cantidad,
                precio_unitario=req_item.precio_unitario,
                notas=req_item.notas  # <--- GUARDA LA NOTA EN CREACIÓN MANUAL
            )
            db.add(nuevo_item)

        db.commit()
        return {"status": "success", "mensaje": "Presupuesto creado con éxito"}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creando presupuesto manual: {str(e)}")


# ==========================================================
#             NUEVAS RUTAS: REGULARIZACIÓN DE PROVEEDORES
# ==========================================================

@app.get("/api/proveedores/pendientes")
def listar_proveedores_pendientes(
        db: Session = Depends(get_db),
        usuario_actual: Usuario = Depends(obtener_usuario_actual)
):
    try:
        from models import Proveedor
        pendientes = db.query(Proveedor).filter(Proveedor.ruc == None).all()

        resultado = []
        for p in pendientes:
            resultado.append({
                "id": p.id,
                "nombre_comercial": p.nombre_comercial or p.razon_social
            })

        return {"status": "success", "data": resultado}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error obteniendo pendientes: {str(e)}")


@app.get("/api/sunat/consultar-ruc/{ruc}")
def consultar_ruc_sunat(
        ruc: str,
        db: Session = Depends(get_db),
        usuario_actual: Usuario = Depends(obtener_usuario_actual)
):
    if len(ruc) != 11:
        raise HTTPException(status_code=400, detail="El RUC debe tener 11 dígitos")

    # Leer el token de forma segura desde las variables de entorno
    TOKEN_DECOLECTA = os.getenv("TOKEN_DECOLECTA")  # <-- 4. Leer desde el .env

    url = f"https://api.decolecta.com/v1/sunat/ruc?numero={ruc}"

    # Pasamos el token por la puerta grande
    headers = {
        "Authorization": f"Bearer {TOKEN_DECOLECTA}",
        "Accept": "application/json"
    }

    try:
        respuesta = requests.get(url, headers=headers)

        if respuesta.status_code == 200:
            datos_sunat = respuesta.json()

            # Verificamos si la respuesta realmente tiene la estructura esperada
            # A veces la API responde 200 pero el contenido dice "error"
            if "nombre" in datos_sunat or "razon_social" in datos_sunat:
                return {"status": "success", "data": datos_sunat}
            else:
                # Caso donde la API responde 200 pero no hay datos útiles
                return {"status": "error", "detail": "La API respondió, pero no se encontró información con ese RUC."}

        else:
            # Aquí capturamos el mensaje real del error que envía el servidor
            error_msg = respuesta.text
            print(f"🔥 ERROR DECOLECTA (Código {respuesta.status_code}): {error_msg}")

            return {
                "status": "error",
                "detail": f"Error de SUNAT ({respuesta.status_code}): El servicio podría estar saturado o el RUC es inválido."
            }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Falla de conexión externa: {str(e)}")

@app.put("/api/proveedores/{id_proveedor}/regularizar")
def regularizar_proveedor(
        id_proveedor: int,
        request: RegularizarProveedorRequest,
        db: Session = Depends(get_db),
        usuario_actual: Usuario = Depends(obtener_usuario_actual)
):
    try:
        from models import Proveedor

        proveedor_db = db.query(Proveedor).filter(Proveedor.id == id_proveedor).first()
        if not proveedor_db:
            raise HTTPException(status_code=404, detail="Proveedor no encontrado")

        proveedor_existente = db.query(Proveedor).filter(Proveedor.ruc == request.ruc).first()
        if proveedor_existente and proveedor_existente.id != id_proveedor:
            raise HTTPException(status_code=400, detail="Ese RUC ya está registrado.")

        proveedor_db.ruc = request.ruc
        proveedor_db.razon_social = request.razon_social.strip().upper()
        proveedor_db.direccion = request.direccion
        proveedor_db.distrito = request.distrito
        proveedor_db.provincia = request.provincia
        proveedor_db.departamento = request.departamento

        db.commit()
        return {"status": "success", "mensaje": "Proveedor regularizado con éxito."}

    except HTTPException as e:
        db.rollback()
        raise e
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error regularizando proveedor: {str(e)}")



# ==========================================================
#             NUEVA RUTA: EXPORTAR A EXCEL (LOGO PIXEL-PERFECT)
# ==========================================================
@app.get("/api/presupuestos/{id_presupuesto}/exportar/excel")
def exportar_excel(
        id_presupuesto: int,
        db: Session = Depends(get_db),
        usuario_actual: Usuario = Depends(obtener_usuario_actual)
):
    try:
        from models import ItemPresupuesto
        import io
        import os
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.drawing.image import Image as ExcelImage
        from fastapi.responses import StreamingResponse

        presupuesto = db.query(Presupuesto).filter(Presupuesto.id == id_presupuesto).first()
        if not presupuesto:
            raise HTTPException(status_code=404, detail="Presupuesto no encontrado")

        items = db.query(ItemPresupuesto).filter(ItemPresupuesto.id_presupuesto == id_presupuesto).all()

        # Configuración de moneda segura
        moneda_db = presupuesto.moneda if getattr(presupuesto, 'moneda', None) else "PEN"
        formato_moneda = '"$" #,##0.00' if moneda_db == 'USD' else (
            '"€" #,##0.00' if moneda_db == 'EUR' else '"S/" #,##0.00')

        borde_fino = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                            bottom=Side(style='thin'))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Cotizacion_{presupuesto.nro_presupuesto or 'SN'}"

        # Cálculo de anchos
        max_desc_len = max([len(str(item.descripcion_original or "")) for item in items] + [45]) if items else 45
        ancho_columna_b = max(45, min(int(max_desc_len * 1.2), 120))

        # ---> SOLUCIÓN: APLICAR EL ANCHO A LAS COLUMNAS <---
        ws.column_dimensions['A'].width = 14  # Ancho para "Ítem" y "Embarcación:"
        ws.column_dimensions['B'].width = ancho_columna_b  # Se expande según el texto/logo
        ws.column_dimensions['C'].width = 10  # Cantidad
        ws.column_dimensions['D'].width = 15  # P. Unitario
        ws.column_dimensions['E'].width = 15  # Subtotal

        # Logo
        ruta_logo = "logo.png"
        if os.path.exists(ruta_logo):
            img = ExcelImage(ruta_logo)
            img.width = int((14 + ancho_columna_b + 40) * 7 + 25)
            img.height = 130
            ws.add_image(img, "A1")

        ws["A8"] = "Embarcación:"
        ws["A8"].font = Font(bold=True)
        ws["B8"] = getattr(presupuesto, 'proyecto_embarcacion', 'Sin especificar')

        ws["A9"] = "Fecha:"
        ws["A9"].font = Font(bold=True)
        ws["B9"] = presupuesto.fecha_emision.strftime("%Y-%m-%d") if getattr(presupuesto, 'fecha_emision',
                                                                             None) else "Sin fecha"

        fila_inicio_tabla = 11

        # NOTA GENERAL (Protegida)
        nota_general = getattr(presupuesto, 'notas', None)
        if nota_general and str(nota_general).strip():
            ws[f"A{fila_inicio_tabla}"] = "Notas:"
            ws[f"A{fila_inicio_tabla}"].font = Font(bold=True, color="FFFF8C00")
            ws[f"B{fila_inicio_tabla}"] = str(nota_general)
            fila_inicio_tabla += 2

        cabeceras = ["Ítem", "Descripción del Trabajo", "Cant.", "P. Unitario", "Subtotal"]
        for col_num, cabecera in enumerate(cabeceras, 1):
            celda = ws.cell(row=fila_inicio_tabla, column=col_num, value=cabecera)
            celda.font = Font(bold=True, color="FFFFFFFF")
            celda.fill = PatternFill(start_color="000F172A", end_color="000F172A", fill_type="solid")
            celda.alignment = Alignment(horizontal="center" if col_num != 2 else "left")
            celda.border = borde_fino

        fila_actual = fila_inicio_tabla + 1
        subtotal_general = 0.0

        for idx, item in enumerate(items, 1):
            ws.cell(row=fila_actual, column=1, value=idx).border = borde_fino

            # NOTA DEL ÍTEM (Protegida)
            texto_desc = str(getattr(item, 'descripcion_original', ''))
            nota_item = getattr(item, 'notas', None)
            if nota_item and str(nota_item).strip():
                texto_desc += f"\n[Obs: {str(nota_item)}]"

            c_desc = ws.cell(row=fila_actual, column=2, value=texto_desc)
            c_desc.alignment = Alignment(wrap_text=True, vertical="center")
            c_desc.border = borde_fino

            cant = float(getattr(item, 'cantidad', 0.0) or 0.0)
            precio = float(getattr(item, 'precio_unitario', 0.0) or 0.0)

            ws.cell(row=fila_actual, column=3, value=cant).border = borde_fino
            c_precio = ws.cell(row=fila_actual, column=4, value=precio)
            c_precio.number_format = formato_moneda
            c_precio.border = borde_fino

            sub_item = cant * precio
            c_sub = ws.cell(row=fila_actual, column=5, value=sub_item)
            c_sub.number_format = formato_moneda
            c_sub.border = borde_fino

            subtotal_general += sub_item
            fila_actual += 1

        # Totales
        igv = subtotal_general * 0.18
        ws.cell(row=fila_actual + 1, column=4, value="SUBTOTAL:").font = Font(bold=True)
        ws.cell(row=fila_actual + 1, column=5, value=subtotal_general).number_format = formato_moneda

        ws.cell(row=fila_actual + 2, column=4, value="IGV (18%):").font = Font(bold=True)
        ws.cell(row=fila_actual + 2, column=5, value=igv).number_format = formato_moneda

        ws.cell(row=fila_actual + 3, column=4, value="TOTAL:").font = Font(bold=True)
        ws.cell(row=fila_actual + 3, column=5, value=subtotal_general + igv).number_format = formato_moneda

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=Cotizacion_{id_presupuesto}.xlsx"}
        )

    except Exception as e:
        import traceback
        traceback.print_exc()  # Esto mostrará el error REAL en tu consola de PyCharm
        raise HTTPException(status_code=500, detail=f"Error en servidor: {str(e)}")